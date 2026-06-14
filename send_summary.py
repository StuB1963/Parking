import smtplib
from email.mime.text import MIMEText
from openpyxl import load_workbook
from datetime import datetime

EXCEL_PATH = "/Users/stuart/Documents/parking_data.xlsx"
ICLOUD_EMAIL = "sbanham1@icloud.com"
ICLOUD_APP_PASSWORD = "Imel-cshu-wąca-rhir"

def generate_summary():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")
    avail = []
    pct = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        timestamp, available, pct_val, *_ = row
        if timestamp.startswith(today):
            avail.append(available)
            pct.append(pct_val)

    if not avail:
        return "No data collected today."

    summary = f"""
Daily Parking Summary — Johnson Street Parkade
Date: {today}

Total Readings: {len(avail)}
Minimum Available: {min(avail)}
Maximum Available: {max(avail)}
Average Available: {sum(avail)/len(avail):.2f}
Average Capacity %: {sum(pct)/len(pct):.2f}%
"""

    return summary

def send_email(body):
    msg = MIMEText(body)
    msg["Subject"] = "Daily Parking Summary — Johnson Street Parkade"
    msg["From"] = ICLOUD_EMAIL
    msg["To"] = ICLOUD_EMAIL

    with smtplib.SMTP("smtp.mail.me.com", 587) as server:
        server.starttls()
        server.login(ICLOUD_EMAIL, ICLOUD_APP_PASSWORD)
        server.send_message(msg)

def main():
    summary = generate_summary()
    send_email(summary)
    print("Summary email sent.")

if __name__ == "__main__":
    main()

