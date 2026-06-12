import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

URL = "https://www.victoria.ca/getting-around/parking/find-parkade-spaces"
EXCEL_PATH = "/Users/stuart/Documents/parking_data.xlsx"
JOHNSON_CAPACITY = 310

def get_johnson_spaces():
    response = requests.get(URL)
    soup = BeautifulSoup(response.text, "html.parser")

    blocks = soup.find_all("div", class_="views-row")

    for block in blocks:
        name_tag = block.find("h3")
        spaces_tag = block.find("div", class_="views-field-field-available-spaces")

        if name_tag and "Johnson" in name_tag.text:
            spaces = spaces_tag.text.strip().replace("Spaces Available", "").strip()
            return int(spaces)

    return None

def append_to_excel(timestamp, spaces, capacity_pct):
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Available Spaces", "Capacity %"])
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([timestamp, spaces, capacity_pct])
    wb.save(EXCEL_PATH)

def main():
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    spaces = get_johnson_spaces()

    if spaces is not None:
        capacity_pct = round((spaces / JOHNSON_CAPACITY) * 100, 2)
        append_to_excel(timestamp, spaces, capacity_pct)
        print(f"Recorded: {spaces} spaces ({capacity_pct}%) at {timestamp}")
    else:
        print("Could not find Johnson Street data.")

if __name__ == "__main__":
    main()
